import asyncio

import openpyxl

from app.processors.base import BaseProcessor


class ExcelProcessor(BaseProcessor):
    async def extract(self, file_path: str) -> list[str]:
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, self._extract_sync, file_path)

    def _extract_sync(self, file_path: str) -> list[str]:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets: list[str] = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                sheets.append(f"[Sheet: {name}]\n" + "\n".join(rows))
        wb.close()
        return sheets if sheets else []
